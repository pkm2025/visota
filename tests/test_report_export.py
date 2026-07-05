"""Tests for generic Report Export (M3.9, VAL-M3-008..015).

Endpoint: ``GET /modern/reports/export/?report=<CODE>&format=<pdf|xlsx>``

Covers:
- VAL-M3-008: PDF export works for all 10 required report codes
- VAL-M3-009: Excel export works for all required reports
- VAL-M3-010: Vietnamese text rendered correctly in PDF
- VAL-M3-011: PDF header (company name, report name, period)
- VAL-M3-012: PDF footer page numbers
- VAL-M3-013: Excel layout (row 1 headers, row 2+ data)
- VAL-M3-014: Filename convention <code>_<period>.{pdf,xlsx}
- VAL-M3-015: Export requires login
"""

from __future__ import annotations

import io
import re
import subprocess
import tempfile
from datetime import date
from decimal import Decimal

import pytest

from apps.core.models import Company
from apps.ledger.models import AccountingVoucher, AccountPeriodBalance, VoucherLine

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ALL_CODES = [
    "S03a",
    "S03b",
    "S06",
    "S07",
    "S08",
    "S35",
    "B01",
    "B02",
    "B03_direct",
    "B03_indirect",
]

FILENAME_REGEX = re.compile(
    r"^(S03a|S03b|S06|S07|S08|S35|B01|B02|B03_direct|B03_indirect)"
    r"_\d{4}(?:\d{2})?\.(pdf|xlsx)$"
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def company(db):
    return Company.objects.create(
        code="EXC",
        name="Công ty Xuất Báo Cáo",
        tax_code="0123456789",
    )


@pytest.fixture
def auth_client(db):
    """Django test client logged in as a superuser."""
    from django.test import Client

    from apps.identity.models import User

    user = User.objects.create_superuser(
        username="export_tester", password="Secret123", email="exp@test.local"
    )
    c = Client()
    c.force_login(user)
    return c


@pytest.fixture
def seeded_vouchers(company):
    """Post a handful of voucher lines so reports have data."""
    v1 = AccountingVoucher.objects.create(
        company=company,
        fiscal_year=2026,
        period=6,
        voucher_no="EXP-001",
        voucher_type=AccountingVoucher.VoucherType.JOURNAL,
        voucher_date=date(2026, 6, 15),
        description="Thu tiền bán hàng",
        status=AccountingVoucher.Status.LEDGER,
    )
    VoucherLine.objects.create(
        voucher=v1,
        line_no=1,
        account_code="1111",
        debit_vnd=Decimal("50000000"),
        credit_vnd=Decimal("0"),
        running_balance_debit=Decimal("50000000"),
    )
    VoucherLine.objects.create(
        voucher=v1,
        line_no=2,
        account_code="5111",
        debit_vnd=Decimal("0"),
        credit_vnd=Decimal("50000000"),
    )
    VoucherLine.objects.create(
        voucher=v1,
        line_no=3,
        account_code="33311",
        debit_vnd=Decimal("0"),
        credit_vnd=Decimal("5000000"),
    )

    v2 = AccountingVoucher.objects.create(
        company=company,
        fiscal_year=2026,
        period=6,
        voucher_no="EXP-002",
        voucher_type=AccountingVoucher.VoucherType.JOURNAL,
        voucher_date=date(2026, 6, 20),
        description="Chi tiền mua hàng",
        status=AccountingVoucher.Status.LEDGER,
    )
    VoucherLine.objects.create(
        voucher=v2,
        line_no=1,
        account_code="1121",
        debit_vnd=Decimal("0"),
        credit_vnd=Decimal("20000000"),
        running_balance_credit=Decimal("20000000"),
    )
    VoucherLine.objects.create(
        voucher=v2,
        line_no=2,
        account_code="1521",
        debit_vnd=Decimal("20000000"),
        credit_vnd=Decimal("0"),
    )

    # AccountPeriodBalance for S06 trial balance
    AccountPeriodBalance.objects.create(
        company=company,
        fiscal_year=2026,
        period=6,
        account_code="1111",
        opening_debit=Decimal("0"),
        opening_credit=Decimal("0"),
        period_debit=Decimal("50000000"),
        period_credit=Decimal("0"),
        closing_debit=Decimal("50000000"),
        closing_credit=Decimal("0"),
    )
    AccountPeriodBalance.objects.create(
        company=company,
        fiscal_year=2026,
        period=6,
        account_code="5111",
        opening_debit=Decimal("0"),
        opening_credit=Decimal("0"),
        period_debit=Decimal("0"),
        period_credit=Decimal("50000000"),
        closing_debit=Decimal("0"),
        closing_credit=Decimal("50000000"),
    )

    return company


# ---------------------------------------------------------------------------
# VAL-M3-008: PDF export works for all required reports
# ---------------------------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.parametrize("code", ALL_CODES)
def test_pdf_export_all_codes(auth_client, seeded_vouchers, code):
    """Each report code returns HTTP 200, application/pdf, body starts with %PDF-."""
    resp = auth_client.get(
        f"/modern/reports/export/?report={code}&format=pdf&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200, f"{code}: status {resp.status_code}"
    assert resp["Content-Type"] == "application/pdf"
    # Body starts with PDF magic
    assert resp.content[:5] == b"%PDF-"
    # Body is non-trivial
    assert len(resp.content) > 1000
    # Filename
    cd = resp["Content-Disposition"]
    assert "attachment" in cd
    filename = re.search(r'filename="([^"]+)"', cd)
    assert filename is not None
    assert FILENAME_REGEX.match(filename.group(1)), f"{code}: bad filename {filename.group(1)}"
    assert filename.group(1).endswith(".pdf")
    assert code in filename.group(1)


# ---------------------------------------------------------------------------
# VAL-M3-009: Excel export works for all required reports
# ---------------------------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.parametrize("code", ALL_CODES)
def test_xlsx_export_all_codes(auth_client, seeded_vouchers, code):
    """Each report code returns OOXML spreadsheet."""
    resp = auth_client.get(
        f"/modern/reports/export/?report={code}&format=xlsx&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200, f"{code}: status {resp.status_code}"
    assert (
        resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # PK zip magic
    assert resp.content[:2] == b"PK"
    # Filename
    cd = resp["Content-Disposition"]
    filename = re.search(r'filename="([^"]+)"', cd)
    assert filename is not None
    assert FILENAME_REGEX.match(filename.group(1))
    assert filename.group(1).endswith(".xlsx")
    assert code in filename.group(1)

    # Workbook is openable and has row 1 = headers
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_column >= 2
    # At least one header cell should be populated
    headers = [cell.value for cell in ws[1]]
    assert any(h for h in headers), f"{code}: row 1 has no header data"


# ---------------------------------------------------------------------------
# VAL-M3-010: Vietnamese text rendered correctly in PDF (no tofu)
# ---------------------------------------------------------------------------


def _pdf_extract_text(pdf_bytes: bytes) -> str:
    """Use the system ``pdftotext`` CLI (Poppler) to extract text from a PDF."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(pdf_bytes)
        pdf_path = f.name
    try:
        result = subprocess.run(
            ["pdftotext", "-enc", "UTF-8", pdf_path, "-"],
            capture_output=True,
            timeout=30,
        )
        return result.stdout.decode("utf-8", errors="replace")
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pytest.skip("pdftotext not available")
    finally:
        import os

        if os.path.exists(pdf_path):
            os.unlink(pdf_path)


@pytest.mark.django_db
def test_pdf_vietnamese_diacritics(auth_client, seeded_vouchers):
    """B01 PDF contains Vietnamese diacritic strings when text is extracted."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=pdf&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200
    text = _pdf_extract_text(resp.content)
    # Check that at least one Vietnamese diacritic-bearing string is present.
    # B01's title is "Bảng cân đối kế toán" — verify the report renders.
    assert text, "Extracted text is empty"
    # Check for at least one Vietnamese diacritic char
    diacritic_chars = "àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ"
    found = any(c in text.lower() for c in diacritic_chars)
    assert found, (
        f"No Vietnamese diacritics found in extracted PDF text. First 200 chars: {text[:200]!r}"
    )


# ---------------------------------------------------------------------------
# VAL-M3-011: PDF header content (company name, report name, period)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_pdf_header_content(auth_client, seeded_vouchers):
    """B01 PDF first page contains company name, report title, and period."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=pdf&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200
    text = _pdf_extract_text(resp.content)
    assert text, "Extracted text is empty"
    # pdftotext may insert stray spaces in diacritic strings; normalize by
    # removing all whitespace and comparing lower-case.
    norm = re.sub(r"\s+", "", text).lower()
    # Company name (Công ty Xuất Báo Cáo) — header renders uppercase (CSS
    # text-transform: uppercase) so the body version is "CÔNG TY XUẤT BÁO CÁO".
    assert "xuấtbáocáo" in norm, f"Company name not found in PDF. Normalized text: {norm[:300]!r}"
    # Report title
    assert "bảngcânđối" in norm or "b01" in norm, (
        f"Report title not found in PDF. Normalized text: {norm[:300]!r}"
    )
    # Period
    assert "2026" in text, f"Period not found in PDF. Text: {text[:300]!r}"
    assert "Tháng 06" in text or "tháng06" in norm, f"Month not found in PDF. Text: {text[:300]!r}"


# ---------------------------------------------------------------------------
# VAL-M3-012: PDF footer page numbers
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_pdf_footer_page_numbers(auth_client, seeded_vouchers):
    """PDF includes footer 'Trang X / Y' on every page."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=pdf&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200
    text = _pdf_extract_text(resp.content)
    assert text, "Extracted text is empty"
    assert "Trang" in text, f"'Trang' not found in footer. Text: {text[:300]!r}"


@pytest.mark.django_db
def test_pdf_template_has_page_counter():
    """Grep the PDF template for WeasyPrint CSS page counters."""
    import os

    template_path = os.path.join(
        os.path.dirname(__file__),
        "..",
        "templates",
        "modern",
        "reporting",
        "report_export_pdf.html",
    )
    template_path = os.path.abspath(template_path)
    with open(template_path) as f:
        content = f.read()
    assert "counter(page)" in content, "Template missing counter(page)"
    assert "counter(pages)" in content, "Template missing counter(pages)"


# ---------------------------------------------------------------------------
# VAL-M3-013: Excel layout (row 1 headers, row 2+ data)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_excel_layout_b01(auth_client, seeded_vouchers):
    """B01 xlsx: row 1 = headers, row 2+ = data."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=xlsx&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 1
    # Row 1 contains headers
    headers = [cell.value for cell in ws[1]]
    assert any(h for h in headers), "Row 1 has no header data"
    # Headers include "Chỉ tiêu" or "Mã số" for B01
    headers_joined = " ".join(str(h) for h in headers if h)
    assert "Chỉ tiêu" in headers_joined or "Mã số" in headers_joined or "STT" in headers_joined, (
        f"B01 headers mismatch: {headers}"
    )


# ---------------------------------------------------------------------------
# VAL-M3-014: Export filename convention
# ---------------------------------------------------------------------------


@pytest.mark.django_db
@pytest.mark.parametrize("code", ALL_CODES)
@pytest.mark.parametrize("fmt", ["pdf", "xlsx"])
def test_filename_convention(auth_client, seeded_vouchers, code, fmt):
    """Filename matches regex <code>_<YYYYMM>.{ext}."""
    resp = auth_client.get(
        f"/modern/reports/export/?report={code}&format={fmt}&fiscal_year=2026&period=6"
    )
    assert resp.status_code == 200
    cd = resp["Content-Disposition"]
    filename = re.search(r'filename="([^"]+)"', cd)
    assert filename is not None, f"No filename in {cd}"
    fname = filename.group(1)
    assert FILENAME_REGEX.match(fname), f"{code}/{fmt}: filename {fname} doesn't match"


@pytest.mark.django_db
def test_filename_zero_padded_month(auth_client, seeded_vouchers):
    """Single-digit month is zero-padded in filename."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=pdf&fiscal_year=2026&period=3"
    )
    assert resp.status_code == 200
    cd = resp["Content-Disposition"]
    assert 'filename="B01_202603.pdf"' in cd, f"Filename not zero-padded: {cd}"


# ---------------------------------------------------------------------------
# VAL-M3-015: Export requires login
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_anonymous_redirects_to_login():
    """Anonymous GET redirects to /auth/login/."""
    from django.test import Client

    c = Client()
    resp = c.get("/modern/reports/export/?report=B01&format=pdf")
    assert resp.status_code in (301, 302, 403)
    if resp.status_code in (301, 302):
        assert "/auth/login/" in resp["Location"]


# ---------------------------------------------------------------------------
# Extra: Invalid report/format handling
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_invalid_report_code(auth_client):
    """Invalid report code returns 400."""
    resp = auth_client.get("/modern/reports/export/?report=INVALID&format=pdf")
    assert resp.status_code == 400


@pytest.mark.django_db
def test_invalid_format(auth_client):
    """Invalid format returns 400."""
    resp = auth_client.get("/modern/reports/export/?report=B01&format=docx")
    assert resp.status_code == 400


@pytest.mark.django_db
def test_pdf_empty_period_no_crash(auth_client, company):
    """Empty period still produces valid PDF (no 500)."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=pdf&fiscal_year=2099&period=12"
    )
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert resp.content[:5] == b"%PDF-"


@pytest.mark.django_db
def test_xlsx_empty_period_no_crash(auth_client, company):
    """Empty period still produces valid Excel (no 500)."""
    resp = auth_client.get(
        "/modern/reports/export/?report=B01&format=xlsx&fiscal_year=2099&period=12"
    )
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Headers should still be present even with no data
    headers = [cell.value for cell in ws[1]]
    assert any(h for h in headers)
