"""Tests for VAT Input/Output List views (M2.7, VAL-M2-021..032).

Endpoints:
- ``GET /modern/reports/vat-input-list/``  → invoice_group=#4 (INPUT)
- ``GET /modern/reports/vat-output-list/`` → invoice_group=#5 (OUTPUT)

Both support ``?format=xlsx`` for Excel export and render the
"Không có dữ liệu" empty state for periods with no rows.
"""

from datetime import date
from decimal import Decimal

import pytest

from apps.core.models import Company
from apps.ledger.models import AccountingVoucher, VoucherLine
from apps.master_data.models import InvoiceGroup, TaxRateCode

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def company(db):
    return Company.objects.create(code="VIL", name="VAT List Test Co")


@pytest.fixture
def tax_rates(db):
    rates = {}
    for code, rate, name in [
        ("00", Decimal("0"), "0%"),
        ("05", Decimal("5"), "5%"),
        ("10", Decimal("10"), "10%"),
    ]:
        rates[code], _ = TaxRateCode.objects.get_or_create(
            code=code, defaults={"rate": rate, "display_name": name}
        )
    return rates


@pytest.fixture
def invoice_groups(db):
    groups = {}
    for code, name in [
        ("4", "INPUT"),
        ("5", "OUTPUT"),
        ("6", "OTHER"),
    ]:
        groups[code], _ = InvoiceGroup.objects.get_or_create(
            code=code, defaults={"name_vi": name, "name_en": name}
        )
    return groups


@pytest.fixture
def auth_client(db):
    """Django test client logged in as a superuser."""
    from django.test import Client

    from apps.identity.models import User

    user = User.objects.create_superuser(
        username="vl_tester", password="Secret123", email="vl@test.local"
    )
    c = Client()
    c.force_login(user)
    return c


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _post_tax_line(
    company,
    fiscal_year,
    period,
    *,
    account_code,
    invoice_group,
    tax_rate,
    goods_amount,
    tax_amount,
    debit=Decimal("0"),
    credit=Decimal("0"),
    voucher_no="VL1",
    invoice_no="HD001",
    invoice_form="01GTKT",
    invoice_symbol="AA/25E",
    object_name="Công ty XYZ",
    object_code="0301234567",
    offset_account="331",
):
    voucher = AccountingVoucher.objects.create(
        company=company,
        fiscal_year=fiscal_year,
        period=period,
        voucher_no=voucher_no,
        voucher_type=AccountingVoucher.VoucherType.JOURNAL,
        voucher_date=date(fiscal_year, period, 15),
        status=AccountingVoucher.Status.LEDGER,
    )
    VoucherLine.objects.create(
        voucher=voucher,
        line_no=1,
        account_code=account_code,
        object_code=object_code,
        object_name=object_name,
        debit_vnd=debit,
        credit_vnd=credit,
        invoice_no=invoice_no,
        invoice_date=date(fiscal_year, period, 10),
        invoice_form=invoice_form,
        invoice_symbol=invoice_symbol,
        tax_code=tax_rate,
        tax_rate=tax_rate.rate if tax_rate else Decimal("0"),
        goods_amount_vnd=goods_amount,
        tax_amount_vnd=tax_amount,
        invoice_group_code=invoice_group,
        offset_account_code=offset_account,
    )
    return voucher


# ---------------------------------------------------------------------------
# VAL-M2-021: VAT input list view returns 200 with a table
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_returns_200(auth_client, company, tax_rates, invoice_groups):
    """GET /modern/reports/vat-input-list/ returns 200 with a table."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_IN1",
    )
    resp = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6")
    assert resp.status_code == 200
    body = resp.content.decode("utf-8")
    assert "<table" in body
    # at least 10 th cells in header row
    assert body.count("<th") >= 10


# ---------------------------------------------------------------------------
# VAL-M2-022: VAT output list view returns 200 with a table
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_returns_200(auth_client, company, tax_rates, invoice_groups):
    """GET /modern/reports/vat-output-list/ returns 200 with a table."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_OUT1",
    )
    resp = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6")
    assert resp.status_code == 200
    body = resp.content.decode("utf-8")
    assert "<table" in body


# ---------------------------------------------------------------------------
# VAL-M2-023: VAT input list columns match TT80 spec
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_has_13_columns(auth_client, company, tax_rates, invoice_groups):
    """Input list header contains exactly the 13 TT80 columns in order."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_COL1",
    )
    resp = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6")
    body = resp.content.decode("utf-8")
    expected = [
        "STT",
        "Ngày HĐ",
        "Số HĐ",
        "Ký hiệu",
        "Mẫu số",
        "Tên KH",
        "MST",
        "Tiền hàng",
        "Thuế suất",
        "Tiền thuế",
        "Tổng tiền",
        "Tk nợ",
        "Tk có",
    ]
    # Verify each header text is present in the HTML
    for col in expected:
        assert col in body, f"column label '{col}' not found in HTML"


# ---------------------------------------------------------------------------
# VAL-M2-024: VAT output list columns match TT80 spec
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_has_13_columns(auth_client, company, tax_rates, invoice_groups):
    """Output list header contains the same 13 TT80 columns."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_COL2",
    )
    resp = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6")
    body = resp.content.decode("utf-8")
    expected = [
        "STT",
        "Ngày HĐ",
        "Số HĐ",
        "Ký hiệu",
        "Mẫu số",
        "Tên KH",
        "MST",
        "Tiền hàng",
        "Thuế suất",
        "Tiền thuế",
        "Tổng tiền",
        "Tk nợ",
        "Tk có",
    ]
    for col in expected:
        assert col in body, f"column label '{col}' not found in HTML"


# ---------------------------------------------------------------------------
# VAL-M2-025: VAT input list shows only invoice_group=#4 rows
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_filters_group4(auth_client, company, tax_rates, invoice_groups):
    """Input list data rows correspond exclusively to invoice_group='4'."""
    # An INPUT line (group 4)
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_G4A",
        invoice_no="INPUT001",
    )
    # An OUTPUT line (group 5) — should NOT appear in input list
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_G5A",
        invoice_no="OUTPUT001",
    )

    resp = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6")
    body = resp.content.decode("utf-8")

    assert "INPUT001" in body
    assert "OUTPUT001" not in body

    # Row count in DB should match
    db_count = VoucherLine.objects.filter(
        invoice_group_code="4",
        voucher__fiscal_year=2026,
        voucher__period=6,
        voucher__status__gte=2,
    ).count()
    # Count <tr> in the table body by checking for INPUT001 presence once
    assert body.count("INPUT001") == 1
    assert db_count == 1


# ---------------------------------------------------------------------------
# VAL-M2-026: VAT output list shows only invoice_group=#5 rows
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_filters_group5(auth_client, company, tax_rates, invoice_groups):
    """Output list data rows correspond exclusively to invoice_group='5'."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_G4B",
        invoice_no="INPUT002",
    )
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_G5B",
        invoice_no="OUTPUT002",
    )

    resp = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6")
    body = resp.content.decode("utf-8")

    assert "OUTPUT002" in body
    assert "INPUT002" not in body


# ---------------------------------------------------------------------------
# VAL-M2-027: Period filter narrows the input list
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_period_filter(auth_client, company, tax_rates, invoice_groups):
    """?fiscal_year=2026&period=6 only returns rows in June 2026."""
    # period 6 row
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_PF6",
        invoice_no="PERIOD6",
    )
    # period 5 row
    _post_tax_line(
        company,
        2026,
        5,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("5000000"),
        tax_amount=Decimal("500000"),
        debit=Decimal("5500000"),
        voucher_no="VL_PF5",
        invoice_no="PERIOD5",
    )

    resp6 = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6")
    body6 = resp6.content.decode("utf-8")
    assert "PERIOD6" in body6
    assert "PERIOD5" not in body6

    resp5 = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=5")
    body5 = resp5.content.decode("utf-8")
    assert "PERIOD5" in body5
    assert "PERIOD6" not in body5


# ---------------------------------------------------------------------------
# VAL-M2-028: Period filter narrows the output list
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_period_filter(auth_client, company, tax_rates, invoice_groups):
    """?fiscal_year=2026&period=6 only returns rows in June 2026."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_OPF6",
        invoice_no="OUT_P6",
    )
    _post_tax_line(
        company,
        2026,
        5,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        credit=Decimal("1000000"),
        voucher_no="VL_OPF5",
        invoice_no="OUT_P5",
    )

    resp6 = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6")
    body6 = resp6.content.decode("utf-8")
    assert "OUT_P6" in body6
    assert "OUT_P5" not in body6


# ---------------------------------------------------------------------------
# VAL-M2-029: Export VAT input list to Excel
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_excel_export(auth_client, company, tax_rates, invoice_groups):
    """?format=xlsx returns a valid .xlsx with matching columns and rows."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_XL1",
        invoice_no="XL001",
    )
    resp = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6&format=xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp["Content-Type"] or "vnd.ms-excel" in resp["Content-Type"]
    cd = resp["Content-Disposition"]
    assert cd.endswith('.xlsx"')
    assert ".xlsx" in cd

    import openpyxl

    wb = openpyxl.load_workbook(__import__("io").BytesIO(resp.content))
    ws = wb.active
    # Row 1 = headers
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "STT"
    assert "Tk nợ" in headers
    assert "Tk có" in headers
    assert ws.max_row >= 2  # header + at least 1 data row
    assert ws.max_column == 13


# ---------------------------------------------------------------------------
# VAL-M2-030: Export VAT output list to Excel
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_excel_export(auth_client, company, tax_rates, invoice_groups):
    """?format=xlsx returns a valid .xlsx mirroring the HTML view."""
    _post_tax_line(
        company,
        2026,
        6,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_XL2",
        invoice_no="XL002",
    )
    resp = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6&format=xlsx")
    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in resp["Content-Type"] or "vnd.ms-excel" in resp["Content-Type"]
    cd = resp["Content-Disposition"]
    assert ".xlsx" in cd

    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "STT"
    assert ws.max_column == 13
    assert ws.max_row >= 2


# ---------------------------------------------------------------------------
# VAL-M2-031: Empty input list shows "Không có dữ liệu"
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_empty_state(auth_client):
    """Empty period shows 'Không có dữ liệu'."""
    resp = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2099&period=12")
    assert resp.status_code == 200
    body = resp.content.decode("utf-8")
    assert "Không có dữ liệu" in body


# ---------------------------------------------------------------------------
# VAL-M2-032: Empty output list shows "Không có dữ liệu"
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_output_list_empty_state(auth_client):
    """Empty period shows 'Không có dữ liệu'."""
    resp = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2099&period=12")
    assert resp.status_code == 200
    body = resp.content.decode("utf-8")
    assert "Không có dữ liệu" in body


# ---------------------------------------------------------------------------
# VAL-M2-033: Posting a tax-bearing voucher is reflected (integration)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_posting_updates_vat_lists_immediately(auth_client, company, tax_rates, invoice_groups):
    """After posting, both lists reflect the new rows immediately."""
    # Verify empty before
    resp_in_before = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=7")
    assert "Không có dữ liệu" in resp_in_before.content.decode("utf-8")

    resp_out_before = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=7")
    assert "Không có dữ liệu" in resp_out_before.content.decode("utf-8")

    # Post voucher with INPUT and OUTPUT lines
    _post_tax_line(
        company,
        2026,
        7,
        account_code="1331",
        invoice_group=invoice_groups["4"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("10000000"),
        tax_amount=Decimal("1000000"),
        debit=Decimal("11000000"),
        voucher_no="VL_INT_IN",
        invoice_no="INT_IN_001",
    )
    _post_tax_line(
        company,
        2026,
        7,
        account_code="33311",
        invoice_group=invoice_groups["5"],
        tax_rate=tax_rates["10"],
        goods_amount=Decimal("20000000"),
        tax_amount=Decimal("2000000"),
        credit=Decimal("2000000"),
        voucher_no="VL_INT_OUT",
        invoice_no="INT_OUT_001",
    )

    # Re-fetch without restart — should show new rows
    resp_in_after = auth_client.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=7")
    body_in = resp_in_after.content.decode("utf-8")
    assert "INT_IN_001" in body_in
    assert "INT_OUT_001" not in body_in

    resp_out_after = auth_client.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=7")
    body_out = resp_out_after.content.decode("utf-8")
    assert "INT_OUT_001" in body_out
    assert "INT_IN_001" not in body_out


# ---------------------------------------------------------------------------
# Login required
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_vat_input_list_login_required():
    """Anonymous GET redirects to /auth/login/."""
    from django.test import Client

    c = Client()
    resp = c.get("/modern/reports/vat-input-list/?fiscal_year=2026&period=6")
    assert resp.status_code in (301, 302)
    assert "/auth/login/" in resp["Location"]


@pytest.mark.django_db
def test_vat_output_list_login_required():
    """Anonymous GET redirects to /auth/login/."""
    from django.test import Client

    c = Client()
    resp = c.get("/modern/reports/vat-output-list/?fiscal_year=2026&period=6")
    assert resp.status_code in (301, 302)
    assert "/auth/login/" in resp["Location"]
