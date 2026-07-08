"""Document text extraction service for the PKM RAG pipeline.

Extracts plain text from common document formats so downstream stages
(chunking, embedding, vector storage) can operate on a uniform ``str`` input.

Supported extensions:
  - ``.pdf``  via :mod:`pypdf`
  - ``.docx`` via :mod:`docx` (python-docx)
  - ``.txt`` / ``.md`` read as UTF-8 text (encoding-tolerant)
  - ``.xlsx`` via :mod:`openpyxl`

Any other extension raises :class:`ValueError`.
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

#: Set of lowercase extensions (without the dot) that ``extract_text`` accepts.
SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({"pdf", "docx", "txt", "md", "xlsx"})


def _detect_extension(file_path: str | os.PathLike[str]) -> str:
    """Return the lowercase extension (no leading dot) for ``file_path``."""
    ext = Path(file_path).suffix.lower().lstrip(".")
    if not ext:
        raise ValueError(
            f"Unsupported file type: {file_path!s} (no extension detected). "
            f"Supported types: {sorted(SUPPORTED_EXTENSIONS)}"
        )
    return ext


def _extract_pdf(file_path: str | os.PathLike[str]) -> str:
    """Extract concatenated page text from a PDF using pypdf."""
    from pypdf import PdfReader

    reader = PdfReader(str(file_path))
    parts: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text:
            parts.append(text)
    return "\n".join(parts).strip()


def _extract_docx(file_path: str | os.PathLike[str]) -> str:
    """Extract paragraph text from a .docx using python-docx."""
    import docx

    document = docx.Document(str(file_path))
    parts: list[str] = [para.text for para in document.paragraphs if para.text]
    return "\n".join(parts).strip()


def _extract_plain_text(file_path: str | os.PathLike[str]) -> str:
    """Read a text file (TXT/MD), decoding gracefully.

    Tries UTF-8 first; on :class:`UnicodeDecodeError` falls back to
    ``utf-8-with-sig`` (handles BOM) then ``latin-1`` (never fails).
    """
    path = Path(file_path)
    encodings = ("utf-8", "utf-8-sig", "latin-1")
    last_error: Exception | None = None
    for enc in encodings:
        try:
            return path.read_text(encoding=enc).strip()
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    # latin-1 cannot fail, but keep the type-checker happy.
    if last_error is not None:  # pragma: no cover - defensive
        raise last_error
    return ""  # pragma: no cover - unreachable


def _extract_xlsx(file_path: str | os.PathLike[str]) -> str:
    """Extract cell text from each sheet of an .xlsx using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cell_texts = [str(cell) for cell in row if cell is not None and str(cell).strip()]
                if cell_texts:
                    parts.append("\t".join(cell_texts))
    finally:
        wb.close()
    return "\n".join(parts).strip()


#: Dispatch table mapping extension -> extraction function.
_EXTRACTORS: dict[str, Any] = {
    "pdf": _extract_pdf,
    "docx": _extract_docx,
    "txt": _extract_plain_text,
    "md": _extract_plain_text,
    "xlsx": _extract_xlsx,
}


def extract_text(file_path: str | os.PathLike[str]) -> str:
    """Extract plain text from a document file.

    Args:
        file_path: Path to the document. The extension determines the parser.

    Returns:
        Extracted text (stripped of surrounding whitespace).

    Raises:
        ValueError: If the file type is not supported.
        FileNotFoundError: If ``file_path`` does not exist.

    The supported formats are: PDF, DOCX, TXT, MD, XLSX.
    """
    ext = _detect_extension(file_path)
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file type '.{ext}'. Supported types: {sorted(SUPPORTED_EXTENSIONS)}"
        )

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    extractor = _EXTRACTORS[ext]
    text: str = extractor(path)
    logger.debug("Extracted %d chars from %s", len(text), file_path)
    return text


__all__ = ["extract_text", "SUPPORTED_EXTENSIONS"]
