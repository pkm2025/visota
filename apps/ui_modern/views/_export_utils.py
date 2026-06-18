"""Shared utilities for Excel (.xlsx) export views."""

from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    OPENPYXL_AVAILABLE = False


def new_workbook(sheet_title: str = "Sheet1"):
    """Create a new openpyxl Workbook with the given first sheet title."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    return wb, ws


def style_header(ws, n_cols: int):
    """Bold + light-blue fill header row."""
    fill = PatternFill("solid", fgColor="4F81BD")
    font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, max_width: int = 60):
    """Autosize columns based on cell content."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(length + 2, 10), max_width)


def xlsx_response(wb, filename: str) -> HttpResponse:
    """Build an HttpResponse that prompts download of `wb` as `filename`."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
