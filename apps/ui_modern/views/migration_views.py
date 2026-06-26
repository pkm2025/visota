"""Migration tool — import Excel data from MISA/Fast/Excel into Visota."""

import io
from decimal import Decimal

from openpyxl import load_workbook

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import redirect, render
from django.views import View

from apps.core.models import Company
from apps.master_data.models import Customer, Vendor, Product


class MigrationUploadView(LoginRequiredMixin, View):
    """Upload Excel file → parse → import customers/vendors/products."""

    template_name = "modern/admin/migration.html"
    login_url = "/auth/login/"

    def get(self, request, *args, **kwargs):
        ctx = {
            "page_title": "Nhập liệu từ Excel",
            "sheets_expected": [
                {"name": "Khách hàng", "columns": "code, name, tax_code, address, phone"},
                {"name": "Nhà cung cấp", "columns": "code, name, tax_code, address, phone"},
                {"name": "Sản phẩm", "columns": "code, name, unit_price, unit, product_type"},
            ],
        }
        return render(request, self.template_name, ctx)

    def post(self, request, *args, **kwargs):
        company = getattr(request, "current_company", None) or Company.objects.first()
        if not company:
            messages.error(request, "Chưa có công ty.")
            return redirect("ui_modern:dashboard")

        uploaded = request.FILES.get("file")
        if not uploaded:
            messages.error(request, "Vui lòng chọn file Excel (.xlsx).")
            return redirect("ui_modern:migration")

        try:
            wb = load_workbook(uploaded, read_only=True, data_only=True)
        except Exception as e:
            messages.error(request, f"Không đọc được file: {e}")
            return redirect("ui_modern:migration")

        stats = {"customers": 0, "vendors": 0, "products": 0, "errors": []}

        # Import Customers (sheet: "Khách hàng" or "Customers" or first sheet)
        customer_sheet = self._find_sheet(wb, ["Khách hàng", "Customers", "khach_hang"])
        if customer_sheet:
            stats["customers"] = self._import_customers(customer_sheet, company, stats["errors"])

        # Import Vendors
        vendor_sheet = self._find_sheet(wb, ["Nhà cung cấp", "Vendors", "nha_cung_cap"])
        if vendor_sheet:
            stats["vendors"] = self._import_vendors(vendor_sheet, company, stats["errors"])

        # Import Products
        product_sheet = self._find_sheet(wb, ["Sản phẩm", "Products", "san_pham"])
        if product_sheet:
            stats["products"] = self._import_products(product_sheet, company, stats["errors"])

        total = stats["customers"] + stats["vendors"] + stats["products"]
        if total == 0:
            messages.warning(request, "Không tìm thấy sheet nào hợp lệ. Kiểm tra lại tên sheet.")
        else:
            msg = f"Đã nhập: {stats['customers']} khách hàng, {stats['vendors']} NCC, {stats['products']} sản phẩm."
            if stats["errors"]:
                msg += f" {len(stats['errors'])} dòng lỗi."
            messages.success(request, msg)

        ctx = {"page_title": "Nhập liệu từ Excel", "result": stats}
        return render(request, self.template_name, ctx)

    def _find_sheet(self, wb, names):
        for name in names:
            if name in wb.sheetnames:
                return wb[name]
        # Try case-insensitive
        for ws in wb.worksheets:
            if ws.title.lower().strip() in [n.lower() for n in names]:
                return ws
        return None

    def _import_customers(self, ws, company, errors):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                code = str(row[0]).strip()
                name = str(row[1]).strip() if len(row) > 1 and row[1] else code
                tax_code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                address = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                phone = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                _, created = Customer.objects.update_or_create(
                    code=code,
                    defaults={
                        "company": company, "name": name,
                        "tax_code": tax_code, "address": address, "phone": phone,
                    },
                )
                if created:
                    count += 1
            except Exception as e:
                errors.append(f"KH dòng {row}: {e}")
        return count

    def _import_vendors(self, ws, company, errors):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                code = str(row[0]).strip()
                name = str(row[1]).strip() if len(row) > 1 and row[1] else code
                tax_code = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                address = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                phone = str(row[4]).strip() if len(row) > 4 and row[4] else ""

                _, created = Vendor.objects.update_or_create(
                    code=code,
                    defaults={
                        "company": company, "name": name,
                        "tax_code": tax_code, "address": address, "phone": phone,
                    },
                )
                if created:
                    count += 1
            except Exception as e:
                errors.append(f"NCC dòng {row}: {e}")
        return count

    def _import_products(self, ws, company, errors):
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                code = str(row[0]).strip()
                name = str(row[1]).strip() if len(row) > 1 and row[1] else code
                price = Decimal(str(row[2])) if len(row) > 2 and row[2] else Decimal("0")
                unit_id = str(row[3]).strip() if len(row) > 3 and row[3] else "cai"
                ptype = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "goods"
                if ptype not in ("goods", "service"):
                    ptype = "goods"

                _, created = Product.objects.update_or_create(
                    code=code,
                    defaults={
                        "company": company, "name": name,
                        "default_unit_price": price, "unit_id": unit_id,
                        "product_type": ptype, "default_vat_rate": Decimal("10"),
                    },
                )
                if created:
                    count += 1
            except Exception as e:
                errors.append(f"SP dòng {row}: {e}")
        return count


class MigrationTemplateView(LoginRequiredMixin, View):
    """Download Excel template for migration."""

    login_url = "/auth/login/"

    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from django.http import HttpResponse

        wb = Workbook()

        # Customers sheet
        ws1 = wb.active
        ws1.title = "Khách hàng"
        ws1.append(["code", "name", "tax_code", "address", "phone"])
        ws1.append(["CUST001", "Công ty ABC", "0101234567", "123 Lê Lợi, HN", "0901234567"])

        # Vendors sheet
        ws2 = wb.create_sheet("Nhà cung cấp")
        ws2.append(["code", "name", "tax_code", "address", "phone"])
        ws2.append(["VENDOR001", "Công ty XYZ", "0109876543", "456 Trần Hưng Đạo, HCM", "0909876543"])

        # Products sheet
        ws3 = wb.create_sheet("Sản phẩm")
        ws3.append(["code", "name", "unit_price", "unit", "product_type"])
        ws3.append(["SP001", "Sản phẩm A", 1000000, "cai", "goods"])
        ws3.append(["SV001", "Dịch vụ B", 5000000, "goi", "service"])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="visota_import_template.xlsx"'
        wb.save(response)
        return response
